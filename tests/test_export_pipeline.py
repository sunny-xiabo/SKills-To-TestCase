import csv
import importlib.util
import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "framework/scripts"


def load_module(name: str, path: Path):
    if str(SCRIPTS) not in sys.path:
        sys.path.insert(0, str(SCRIPTS))
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


COMMON = load_module("testcase_common", SCRIPTS / "testcase_common.py")


SAMPLE_MD = """## 用户中心

| 用例ID | 所属模块 | 测试点 | 前置条件 | 操作步骤 | 预期结果 | 关联检查点 | 场景类型 | 优先级 | 备注 |
|--------|----------|--------|----------|----------|----------|------------|----------|--------|------|
| TC-001 | 用户中心 | 保存资料 | 已登录 | 1. 输入姓名 2. 单击保存按钮 | 1. 提示“保存成功” | UC-01 | 正向 | P1 | |
| TC-002 | 用户中心 | 并发提交 | 已登录 | 1. 同时提交两次 | 1. 仅成功一次 | RISK-01 | 并发 |  | |
"""


class ExportPipelineTest(unittest.TestCase):
    def test_priority_rules_align_with_docs(self):
        self.assertEqual("P2", COMMON.default_priority_for_type("并发"))
        self.assertEqual("Low", COMMON.priority_to_jira("", "并发"))
        self.assertEqual("High", COMMON.priority_to_jira("", "异常"))
        self.assertEqual("Medium", COMMON.priority_to_jira("P1", "并发"))

    def test_md_to_json_and_csv_share_parser(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            md_path = temp_path / "2-用例定稿.md"
            json_path = temp_path / "export_data.json"
            csv_path = temp_path / "jira_export.csv"
            md_path.write_text(SAMPLE_MD, encoding="utf-8")

            json_result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "md_to_json.py"),
                    str(md_path),
                    str(json_path),
                    "--project",
                    "演示项目",
                    "--module",
                    "用户中心",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(0, json_result.returncode, json_result.stderr)
            payload = json.loads(json_path.read_text(encoding="utf-8"))
            self.assertEqual(2, len(payload["testcases"]))
            concurrent = next(tc for tc in payload["testcases"] if tc["id"] == "TC-002")
            self.assertEqual("P2", concurrent["priority"])
            self.assertEqual("演示项目", payload["meta"]["project"])

            csv_result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "md_to_csv.py"),
                    str(md_path),
                    str(csv_path),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(0, csv_result.returncode, csv_result.stderr)
            with csv_path.open(encoding="utf-8-sig") as handle:
                rows = list(csv.reader(handle))
            # 表头 + 至少两条用例首行
            self.assertGreaterEqual(len(rows), 3)
            concurrent_rows = [row for row in rows if row and row[0] == "TC-002"]
            self.assertEqual(1, len(concurrent_rows))
            self.assertEqual("Low", concurrent_rows[0][3])

    def test_case_sort_key_orders_suffix(self):
        ids = ["TC-002", "TC-001a", "TC-001", "TC-010"]
        ordered = sorted(ids, key=COMMON.case_sort_key)
        self.assertEqual(["TC-001", "TC-001a", "TC-002", "TC-010"], ordered)


class SkillContractTest(unittest.TestCase):
    def test_creator_routes_change_and_export_md_pipeline(self):
        skill_dir = ROOT / "skills/testcase-creator"
        prompt = (skill_dir / "prompt.md").read_text(encoding="utf-8")
        export_ref = (skill_dir / "references/export-workflow.md").read_text(encoding="utf-8")
        change_ref = skill_dir / "references/change-workflow.md"
        input_ref = (skill_dir / "references/input-and-generation.md").read_text(encoding="utf-8")

        self.assertTrue(change_ref.is_file())
        self.assertIn("references/testcase-creator/change-workflow.md", prompt)
        self.assertIn("增量变更", prompt)
        self.assertIn("md_to_json.py", prompt)
        self.assertIn("md_to_json.py", export_ref)
        self.assertIn("采用推荐", input_ref)
        self.assertIn("历史用例复用", input_ref)
        self.assertIn("禁止手写", (ROOT / "skills/testcase-export/prompt.md").read_text(encoding="utf-8"))

    def test_versions(self):
        creator = (ROOT / "skills/testcase-creator/meta.yaml").read_text(encoding="utf-8")
        export = (ROOT / "skills/testcase-export/meta.yaml").read_text(encoding="utf-8")
        self.assertIn('version: "1.12.0"', creator)
        self.assertIn('version: "1.9.0"', export)
        versions = load_module("framework_versions", SCRIPTS / "framework_versions.py")
        self.assertEqual("1.12.0", versions.EXPECTED["testcase-creator"])
        self.assertEqual("1.9.0", versions.EXPECTED["testcase-export"])

    def test_prompt_requires_version_check_and_merge_script(self):
        prompt = (ROOT / "skills/testcase-creator/prompt.md").read_text(encoding="utf-8")
        change = (
            ROOT / "skills/testcase-creator/references/change-workflow.md"
        ).read_text(encoding="utf-8")
        export_ref = (
            ROOT / "skills/testcase-creator/references/export-workflow.md"
        ).read_text(encoding="utf-8")
        self.assertIn("check_environment.py", prompt)
        self.assertIn("gate_stage.py", prompt)
        self.assertIn("recommend_checkpoints.py", prompt)
        self.assertIn("scan_code_scope.py", prompt)
        self.assertIn("A–I", prompt)
        self.assertIn("export_all.py", export_ref)
        self.assertIn("merge_cases.py", change)
        self.assertIn("gate_stage.py", change)
        self.assertIn("suggest_assets_from_bugs.py", prompt)
        input_ref = (
            ROOT / "skills/testcase-creator/references/input-and-generation.md"
        ).read_text(encoding="utf-8")
        self.assertIn("来源 I：代码", input_ref)
        self.assertIn("需求与代码差异", input_ref)


class ExcelPresentationTest(unittest.TestCase):
    def test_normalize_and_export_status_dropdown(self):
        excel = load_module("export_excel", SCRIPTS / "export_excel.py")
        self.assertEqual(
            "1. 打开页面\n2. 单击保存",
            excel.normalize_multiline("1、打开页面\n\n2.单击保存\n"),
        )
        self.assertEqual(
            "1. a\n2. b",
            excel.normalize_multiline("1. a 2. b"),
        )
        height = excel.estimate_row_height(
            "1. " + ("步骤文字" * 20),
            "1. 预期",
            "前置",
            "测试点",
        )
        self.assertGreaterEqual(height, excel.MIN_ROW_HEIGHT)
        self.assertLessEqual(height, excel.MAX_ROW_HEIGHT)

        payload = {
            "meta": {
                "project": "演示",
                "module": "模块A",
                "generated_at": "2026-08-14",
                "author": "测试同学",
            },
            "testcases": [
                {
                    "id": "TC-001",
                    "module": "模块A",
                    "test_point": "保存",
                    "precondition": "已登录",
                    "steps": "1. 输入 2. 保存",
                    "expected": "1. 成功",
                    "checkpoint": "UC-01",
                    "type": "正向",
                    "priority": "P1",
                    "remark": "",
                },
                {
                    "id": "TC-002",
                    "module": "模块B",
                    "test_point": "异常",
                    "precondition": "",
                    "steps": "1. 提交",
                    "expected": "1. 提示错误",
                    "checkpoint": "",
                    "type": "异常",
                    "priority": "P0",
                    "remark": "保留备注",
                },
            ],
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            assets = temp_path / ".testcase-assets"
            assets.mkdir()
            (assets / "project.config.md").write_text(
                "| 测试负责人 | 配置负责人 |\n", encoding="utf-8"
            )
            history = assets / "history" / "run1"
            history.mkdir(parents=True)
            input_path = history / "export_data.json"
            output_path = history / "testcases.xlsx"
            input_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "export_excel.py"),
                    str(input_path),
                    str(output_path),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(0, result.returncode, result.stderr + result.stdout)
            self.assertTrue(output_path.is_file())

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            ws = wb["测试用例"]
            self.assertIn("项目：演示", ws["A1"].value)
            self.assertEqual("未执行", ws.cell(row=3, column=10).value)
            self.assertEqual("测试同学", ws.cell(row=3, column=11).value)
            self.assertEqual("1. 输入\n2. 保存", ws.cell(row=3, column=5).value)
            self.assertTrue(ws.data_validations.dataValidation)
            self.assertIn("统计", wb.sheetnames)
            stat = wb["统计"]
            titles = [stat.cell(row=r, column=1).value for r in range(1, 40)]
            self.assertTrue(any(t and "优先级" in str(t) for t in titles))
            self.assertTrue(any(t and "所属模块" in str(t) for t in titles))


class InitSyncTest(unittest.TestCase):
    def test_init_script_documents_sync_mode(self):
        sh = (ROOT / "init-testcase.sh").read_text(encoding="utf-8")
        ps1 = (ROOT / "init-testcase.ps1").read_text(encoding="utf-8")
        self.assertIn("--sync", sh)
        self.assertIn("项目资产已保护", sh)
        self.assertIn("FRAMEWORK_VERSION", sh)
        self.assertIn("md_to_json.py", sh)
        self.assertIn("export_all.py", sh)
        self.assertIn("-Sync", ps1)
        self.assertIn("FRAMEWORK_VERSION", ps1)
        self.assertTrue((ROOT / "sync-projects.sh").is_file())


class VersionAndMergeTest(unittest.TestCase):
    def test_framework_version_check(self):
        versions = load_module("framework_versions", SCRIPTS / "framework_versions.py")
        with tempfile.TemporaryDirectory() as temp_dir:
            assets = Path(temp_dir) / ".testcase-assets"
            assets.mkdir()
            ok, messages = versions.check_versions(None)
            self.assertFalse(ok)

            path = versions.write_version_file(assets)
            ok, messages = versions.check_versions(path)
            self.assertTrue(ok, messages)

            path.write_text("testcase-creator=0.0.1\ntestcase-export=0.0.1\n", encoding="utf-8")
            ok, messages = versions.check_versions(path)
            self.assertFalse(ok)

    def test_merge_cases_script(self):
        baseline = """## 模块
| 用例ID | 所属模块 | 测试点 | 前置条件 | 操作步骤 | 预期结果 | 关联检查点 | 场景类型 | 优先级 | 备注 |
|--------|----------|--------|----------|----------|----------|------------|----------|--------|------|
| TC-001 | 模块 | 旧点 | 无 | 1. a | 1. ok | UC-01 | 正向 | P1 | |
| TC-002 | 模块 | 待废 | 无 | 1. b | 1. ok | | 异常 | P0 | |
"""
        changeset = """## 变更集
### 新增
| 用例ID | 所属模块 | 测试点 | 前置条件 | 操作步骤 | 预期结果 | 关联检查点 | 场景类型 | 优先级 | 备注 |
|--------|----------|--------|----------|----------|----------|------------|----------|--------|------|
| TC-003 | 模块 | 新点 | 无 | 1. c | 1. ok | | 边界 | P1 | |

### 修改
| 用例ID | 所属模块 | 测试点 | 前置条件 | 操作步骤 | 预期结果 | 关联检查点 | 场景类型 | 优先级 | 备注 |
|--------|----------|--------|----------|----------|----------|------------|----------|--------|------|
| TC-001 | 模块 | 新描述 | 无 | 1. a2 | 1. ok2 | UC-01 | 正向 | P1 | |

### 废弃
| 用例ID | 测试点 | 废弃原因 |
|--------|--------|----------|
| TC-002 | 待废 | 需求删除 |
"""
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            base = temp_path / "base.md"
            change = temp_path / "change.md"
            out = temp_path / "merged.md"
            base.write_text(baseline, encoding="utf-8")
            change.write_text(changeset, encoding="utf-8")
            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "merge_cases.py"),
                    "--baseline",
                    str(base),
                    "--changeset",
                    str(change),
                    "--output",
                    str(out),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(0, result.returncode, result.stderr + result.stdout)
            text = out.read_text(encoding="utf-8")
            self.assertIn("TC-001", text)
            self.assertIn("新描述", text)
            self.assertIn("TC-003", text)
            self.assertNotIn("TC-002", text.split("## 模块")[-1] if "## 模块" in text else text)

    def test_export_all_priority_filter_and_csv_tools(self):
        md = """## 用户中心
| 用例ID | 所属模块 | 测试点 | 前置条件 | 操作步骤 | 预期结果 | 关联检查点 | 场景类型 | 优先级 | 备注 |
|--------|----------|--------|----------|----------|----------|------------|----------|--------|------|
| TC-001 | 用户中心 | 保存 | 已登录 | 1. 保存 | 1. 成功 | UC-01 | 正向 | P1 | |
| TC-002 | 用户中心 | 崩溃 | 无 | 1. 提交 | 1. 提示 | | 异常 | P0 | |
| TC-003 | 用户中心 | 并发 | 无 | 1. 双击 | 1. 一次 | | 并发 | P2 | |
"""
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            md_path = temp_path / "2-用例定稿.md"
            md_path.write_text(md, encoding="utf-8")
            out_dir = temp_path / "out"
            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "export_all.py"),
                    str(md_path),
                    "--out-dir",
                    str(out_dir),
                    "--formats",
                    "j,e",
                    "--priority",
                    "P0,P1",
                    "--project",
                    "演示",
                    "--module",
                    "用户中心",
                    "--skip-quality",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(0, result.returncode, result.stderr + result.stdout)
            payload = json.loads((out_dir / "export_data.json").read_text(encoding="utf-8"))
            self.assertEqual(2, len(payload["testcases"]))
            self.assertTrue((out_dir / "testcases-smoke.xlsx").is_file())
            self.assertTrue((out_dir / "jira_export-smoke.csv").is_file())

            tapd = temp_path / "tapd.csv"
            csv_result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "md_to_csv.py"),
                    str(md_path),
                    str(tapd),
                    "--tool",
                    "tapd",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(0, csv_result.returncode, csv_result.stderr)
            self.assertIn("用例名称", tapd.read_text(encoding="utf-8-sig"))

    def test_suggest_assets_from_bugs(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            bugs = temp_path / "bugs.txt"
            bugs.write_text("BUG-100\t提交重复数据\n空指针崩溃\n", encoding="utf-8")
            out = temp_path / "cand.md"
            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "suggest_assets_from_bugs.py"),
                    str(bugs),
                    "--kind",
                    "both",
                    "--output",
                    str(out),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(0, result.returncode, result.stderr + result.stdout)
            text = out.read_text(encoding="utf-8")
            self.assertIn("BUG-01", text)
            self.assertIn("提交重复数据", text)

    def test_environment_gate_and_recommend(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            assets = temp_path / ".testcase-assets"
            assets.mkdir()
            (assets / "checkpoints-index.md").write_text(
                "## 列表\n- [LIST-01] 分页与筛选\n- [FILE-01] 附件上传\n- [RISK-01] 权限校验 [已废弃]\n",
                encoding="utf-8",
            )
            (assets / "review-expectations-index.md").write_text("# r\n", encoding="utf-8")
            versions = load_module("framework_versions", SCRIPTS / "framework_versions.py")
            versions.write_version_file(assets)

            env = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "check_environment.py"),
                    str(temp_path),
                    "--strict",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(0, env.returncode, env.stdout + env.stderr)
            self.assertIn("[OK]", env.stdout)

            gate = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "gate_stage.py"),
                    "--stage",
                    "init",
                    "--path",
                    str(temp_path),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(0, gate.returncode, gate.stdout + gate.stderr)
            self.assertIn("[GATE OK] stage=init", gate.stdout)

            rec = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "recommend_checkpoints.py"),
                    "--checkpoints",
                    str(assets / "checkpoints-index.md"),
                    "--text",
                    "列表支持分页，需要上传附件",
                    "--output",
                    str(temp_path / "rec.md"),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(0, rec.returncode, rec.stdout + rec.stderr)
            rec_text = (temp_path / "rec.md").read_text(encoding="utf-8")
            self.assertIn("LIST-01", rec_text)
            self.assertIn("FILE-01", rec_text)
            self.assertNotIn("RISK-01", rec_text)

            run_dir = assets / "history" / "run1"
            run_dir.mkdir(parents=True)
            # merge gate fail without files
            g_fail = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "gate_stage.py"),
                    "--stage",
                    "merge",
                    "--run-dir",
                    str(run_dir),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(1, g_fail.returncode)
            self.assertIn("[GATE FAIL]", g_fail.stdout)

            (run_dir / "1-变更集.md").write_text(
                "### 新增\n### 修改\n### 废弃\n", encoding="utf-8"
            )
            (run_dir / "1-评审记要.md").write_text(
                "## 变更合并摘要\n- 合并后有效：1 条\n\n## 模块\n"
                "| 用例ID | 所属模块 | 测试点 | 前置条件 | 操作步骤 | 预期结果 | 关联检查点 | 场景类型 | 优先级 |\n"
                "|--------|----------|--------|----------|----------|----------|------------|----------|--------|\n"
                "| TC-001 | 模块 | 点 | 无 | 1. a | 1. b | | 正向 | P1 |\n",
                encoding="utf-8",
            )
            g_ok = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "gate_stage.py"),
                    "--stage",
                    "merge",
                    "--run-dir",
                    str(run_dir),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(0, g_ok.returncode, g_ok.stdout + g_ok.stderr)

            # history two-level
            (run_dir / "2-用例定稿.md").write_text(
                "## 组织树\n"
                "| 用例ID | 所属模块 | 测试点 | 前置条件 | 操作步骤 | 预期结果 | 关联检查点 | 场景类型 | 优先级 |\n"
                "|--------|----------|--------|----------|----------|----------|------------|----------|--------|\n"
                "| TC-009 | 组织树 | 展开节点 | 无 | 1. 点 | 1. 展开 | | 正向 | P1 |\n",
                encoding="utf-8",
            )
            # rename dir for alias match
            org_dir = assets / "history" / "20260101_组织树演示"
            run_dir.rename(org_dir)
            hist = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "recommend_history.py"),
                    "--history-root",
                    str(assets / "history"),
                    "--module",
                    "组织树",
                    "--list-dirs",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(0, hist.returncode, hist.stdout + hist.stderr)
            self.assertIn("组织树", hist.stdout)

    def test_scan_code_scope_requires_range_and_lists_files(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            src = temp_path / "src"
            src.mkdir()
            (src / "login.py").write_text(
                "def login(user, token):\n"
                "    if not token:\n"
                "        raise PermissionError('401')\n"
                "    return {'ok': True}\n",
                encoding="utf-8",
            )
            out = temp_path / "scope.md"
            # no range -> fail
            bad = subprocess.run(
                [sys.executable, str(SCRIPTS / "scan_code_scope.py")],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(2, bad.returncode)

            ok = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPTS / "scan_code_scope.py"),
                    "--path",
                    str(src),
                    "--output",
                    str(out),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(0, ok.returncode, ok.stdout + ok.stderr)
            text = out.read_text(encoding="utf-8")
            self.assertIn("login.py", text)
            self.assertIn("权限/鉴权", text)
            self.assertIn("不是用例", text)


if __name__ == "__main__":
    unittest.main()
